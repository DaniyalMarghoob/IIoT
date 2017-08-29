
from openpyxl import Workbook
from openpyxl import load_workbook
from gpiozero import CPUTemperature
import os.path
import xlsxwriter
import sys
import os
import glob
import time

path='/home/pi/Thesis/MQTTPythonClient/retainData.xlsx'

def createbook():
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet('data')
    worksheet.write(0, 0, 0)
    workbook.close()
    print("new file created")

def readData():
    if(not os.path.isfile(path)):
        createbook()

    wb = load_workbook(filename=path)
    sheet=wb['data']
    data=[]
    for i in range (1,sheet.max_row+1):
        data.append(sheet.cell(row=i,column=1).value)
    return data



def writeData(data):
    wb = Workbook()
    destFile = path
    sheet1 = wb.active
    sheet1.title = 'data'
    for i in range(1, len(data)+1):
        _ = sheet1.cell(column=1, row=i, value=data[i-1])
    wb.save(filename=destFile)
    wb.close()

def cpuTemp():
	cpu=CPUTemperature()
	return cpu.temperature


  
   
def readTempRawMachine():
    base_dir='/sys/bus/w1/devices/'
    device_folder=glob.glob(base_dir+'28*')[0]
    device_file=device_folder+'/w1_slave'
    f=open(device_file,'r')
    lines=f.readlines()
    f.close()
    return lines

def machineTemperature():
    lines=readTempRawMachine()
    while lines[0].strip()[-3:]!='YES':
        time.sleep(0.2)
        lines=readTempRaw()
    equals_pos=lines[1].find('t=')
    if equals_pos !=1:
        tempString=lines[1][equals_pos+2:]
        temp_c=float(tempString)/1000
        temp_f=temp_c * (9.0/(5.0+32.0))
        return temp_c
                     
def readTempRawEnv():
    base_dir='/sys/bus/w1/devices/'
    device_folder=glob.glob(base_dir+'28*')[1]
    device_file=device_folder+'/w1_slave'
    f=open(device_file,'r')
    lines=f.readlines()
    f.close()
    return lines

def envTemperature():
    lines=readTempRawEnv()
    while lines[0].strip()[-3:]!='YES':
        time.sleep(0.2)
        lines=readTempRaw()
    equals_pos=lines[1].find('t=')
    if equals_pos !=1:
        tempString=lines[1][equals_pos+2:]
        temp_c=float(tempString)/1000
        temp_f=temp_c * (9.0/(5.0+32.0))
        return temp_c                     